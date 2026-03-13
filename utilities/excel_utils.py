import openpyxl


def getRowCount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_row


def getColumnCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_column


def readData(file,sheetname,rownum,columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.cell(rownum,columnnum).value


def WriteData(file,sheetname,rownum,columnnum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(rownum, columnnum).value=data
    workbook.save(file)


